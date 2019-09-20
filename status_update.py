#update_status
#Trace Watkins
#8/7/2019

import regex
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.alignment import Alignment
import regex as re


def excel_write(sheet,access_list):
    #find column with IP addresses
    for row in sheet.iter_rows(min_col=1, max_row=274, max_col=1):
        for cell in row:
            status_check = re_status(cell.value)
            print(cell.value, status_check)
            if status_check == "ip reply/reachable":
                access_list.append(cell.value)
            cell_row = cell.row
            sheet.cell(row = cell_row, column = 2).value = status_check
        

def re_status(device):
    #Create files in C:\Users\watkinst\print_html with dict of ip and status
    if isinstance(device, str):
        f = open("U:\\Python\\Python37-32\\office_scan\pings\\" + device + ".txt")
        status_doc = f.read()
        reached = re.search('bytes',status_doc)
        if reached == None:
            unreachable = re.search('unreachable',status_doc)
            if unreachable == None:
                time_out = re.search ('request',status_doc)
                return('no ip reply')
            else:
                return ('ip reply/unreacheable')
        else:
            return('ip reply/reachable')
            
            
            

#load devices workbook
access_list = []
wb = load_workbook(filename ="U:\Python\Python37-32\office_scan\Aug 2019_OfficeScan Compliance Report.xlsx")
comps = wb['devices']
excel_write(comps,access_list)
wb.save(filename ="U:\Python\Python37-32\office_scan\Aug 2019_OfficeScan Compliance Report.xlsx")
print (access_list)
